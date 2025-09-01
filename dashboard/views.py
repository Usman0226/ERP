from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import connection
from django.apps import apps
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from accounts.models import User, Role, Permission, AuthIdentifier, UserSession, AuditLog, FailedLogin
from students.models import Student, StudentEnrollmentHistory, StudentDocument, CustomField, StudentImport
from faculty.models import Faculty, FacultySubject, FacultySchedule, FacultyLeave, FacultyPerformance, FacultyDocument, CustomField as FacultyCustomField, CustomFieldValue
from django.utils import timezone
from django.db.models import Avg
from django.http import HttpResponse
from django.db.models import Q, Count
import csv
import os

from .models import APICollection, APIEnvironment, APIRequest, APITest, APITestResult, APITestSuite, APITestSuiteResult, APIAutomation


def is_admin(user):
    return user.is_authenticated and user.is_staff


@login_required
@user_passes_test(is_admin)
def dashboard_home(request):
    """Main dashboard view with statistics"""
    context = {
        'total_users': User.objects.count(),
        'total_roles': Role.objects.count(),
        'total_permissions': Permission.objects.count(),
        'active_sessions': UserSession.objects.filter(revoked=False).count(),
        'total_students': Student.objects.count(),
        'active_students': Student.objects.filter(status='ACTIVE').count(),
        'total_faculty': Faculty.objects.count(),
        'active_faculty': Faculty.objects.filter(status='ACTIVE', currently_associated=True).count(),
        'total_custom_fields': CustomField.objects.filter(is_active=True).count(),
        'total_faculty_custom_fields': FacultyCustomField.objects.filter(is_active=True).count(),
        'recent_users': User.objects.order_by('-date_joined')[:10],
        'recent_logins': AuditLog.objects.filter(action='login').order_by('-created_at')[:10],
        'recent_students': Student.objects.order_by('-created_at')[:5],
        'recent_faculty': Faculty.objects.order_by('-created_at')[:5],
    }
    return render(request, 'dashboard/home.html', context)


@login_required
@user_passes_test(is_admin)
def users_list(request):
    """Users management page"""
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'dashboard/users.html', {'users': users})


@login_required
@user_passes_test(is_admin)
def roles_list(request):
    """Roles management page"""
    roles = Role.objects.all().order_by('name')
    return render(request, 'dashboard/roles.html', {'roles': roles})


@login_required
@user_passes_test(is_admin)
def sessions_list(request):
    """Active sessions page"""
    sessions = UserSession.objects.all().order_by('-created_at')
    return render(request, 'dashboard/sessions.html', {'sessions': sessions})


@login_required
@user_passes_test(is_admin)
def audit_logs(request):
    """Audit logs page"""
    logs = AuditLog.objects.all().order_by('-created_at')[:100]
    return render(request, 'dashboard/audit_logs.html', {'logs': logs})


@login_required
@user_passes_test(is_admin)
def database_schema(request):
    """Database schema overview"""
    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [row[0] for row in cursor.fetchall()]
    
    schema_info = []
    for table in tables:
        with connection.cursor() as cursor:
            cursor.execute(f"PRAGMA table_info({table});")
            columns = cursor.fetchall()
            schema_info.append({
                'table': table,
                'columns': columns
            })
    
    return render(request, 'dashboard/schema.html', {'schema_info': schema_info})


# API Endpoints for dashboard data
@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_database_schema(request):
    """API endpoint to get database schema information"""
    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [row[0] for row in cursor.fetchall()]
    
    schema_data = {}
    for table in tables:
        with connection.cursor() as cursor:
            cursor.execute(f"PRAGMA table_info({table});")
            columns = cursor.fetchall()
            schema_data[table] = [
                {
                    'id': col[0],
                    'name': col[1],
                    'type': col[2],
                    'not_null': bool(col[3]),
                    'default': col[4],
                    'primary_key': bool(col[5])
                }
                for col in columns
            ]
    
    return Response(schema_data)


@login_required
@user_passes_test(is_admin)
def download_schema_excel(request):
    """Download database schema as an Excel file with one sheet per table."""
    # Lazy import to avoid hard dependency if not used
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("openpyxl is not installed. Please install it to use Excel export.", status=500)
    except Exception as exc:
        return HttpResponse(f"Error importing openpyxl: {exc}", status=500)

    wb = Workbook()
    wb.remove(wb.active)

    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [row[0] for row in cursor.fetchall()]

    for table in tables:
        ws = wb.create_sheet(title=str(table)[:31])  # Excel sheet name limit
        ws.append(["#", "column", "type", "not_null", "default", "primary_key"])  
        with connection.cursor() as cursor:
            cursor.execute(f"PRAGMA table_info({table});")
            for col in cursor.fetchall():
                ws.append([col[0], col[1], col[2], bool(col[3]), col[4], bool(col[5])])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="database_schema.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def download_schema_excel_single(request):
    """Download database schema as a single-sheet Excel file consolidating all tables."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("openpyxl is not installed. Please install it to use Excel export.", status=500)
    except Exception as exc:
        return HttpResponse(f"Error importing openpyxl: {exc}", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = 'schema'
    ws.append(["table", "#", "column", "type", "not_null", "default", "primary_key"])  

    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [row[0] for row in cursor.fetchall()]

    for table in tables:
        with connection.cursor() as cursor:
            cursor.execute(f"PRAGMA table_info({table});")
            for col in cursor.fetchall():
                ws.append([table, col[0], col[1], col[2], bool(col[3]), col[4], bool(col[5])])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="database_schema_single.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def download_schema_csv(request):
    """Download database schema as CSV with all tables consolidated."""
    try:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="database_schema.csv"'

        writer = csv.writer(response)
        writer.writerow(["table", "#", "column", "type", "not_null", "default", "primary_key"])  

        with connection.cursor() as cursor:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [row[0] for row in cursor.fetchall()]

        for table in tables:
            with connection.cursor() as cursor:
                cursor.execute(f"PRAGMA table_info({table});")
                for col in cursor.fetchall():
                    writer.writerow([table, col[0], col[1], col[2], bool(col[3]), col[4], bool(col[5])])

        return response
    except Exception as e:
        return HttpResponse(f"Error generating CSV: {e}", status=500)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_table_data(request, table_name):
    """API endpoint to get data from any table"""
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT * FROM {table_name} LIMIT 100;")
            columns = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()
            
            data = []
            for row in rows:
                data.append(dict(zip(columns, row)))
        
        return Response({
            'table': table_name,
            'columns': columns,
            'data': data,
            'count': len(data)
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_dashboard_stats(request):
    """API endpoint for dashboard statistics"""
    stats = {
        'users': {
            'total': User.objects.count(),
            'active': User.objects.filter(is_active=True).count(),
            'staff': User.objects.filter(is_staff=True).count(),
            'verified': User.objects.filter(is_verified=True).count(),
        },
        'auth': {
            'roles': Role.objects.count(),
            'permissions': Permission.objects.count(),
            'identifiers': AuthIdentifier.objects.count(),
            'active_sessions': UserSession.objects.filter(revoked=False).count(),
        },
        'security': {
            'failed_logins': FailedLogin.objects.count(),
            'audit_logs': AuditLog.objects.count(),
        }
    }
    return Response(stats)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_models_info(request):
    """API endpoint to get information about all Django models"""
    models_info = {}
    
    for model in apps.get_models():
        app_label = model._meta.app_label
        model_name = model._meta.model_name
        
        if app_label not in models_info:
            models_info[app_label] = {}
        
        fields_info = []
        for field in model._meta.fields:
            fields_info.append({
                'name': field.name,
                'type': field.__class__.__name__,
                'null': field.null,
                'blank': field.blank,
                'unique': field.unique,
            })
        
        models_info[app_label][model_name] = {
            'table_name': model._meta.db_table,
            'fields': fields_info,
            'count': model.objects.count() if hasattr(model.objects, 'count') else 0,
        }
    
    return Response(models_info)


# ------------------------------
# ER diagram generation
# ------------------------------
def _generate_mermaid_er_diagram() -> str:
    """Build Mermaid ER diagram from Django model metadata."""
    # Header
    lines = ["erDiagram"]

    # Collect nodes
    all_models = list(apps.get_models())
    for model in all_models:
        model_label = model.__name__
        lines.append(f"    {model_label} {{")
        for field in model._meta.get_fields():
            # Skip reverse relations rendered later via FK fields
            if hasattr(field, 'remote_field') and field.auto_created and not field.concrete:
                continue
            if getattr(field, 'many_to_many', False):
                continue
            # Field name and type
            field_name = getattr(field, 'name', 'id')
            field_type = getattr(field, 'get_internal_type', lambda: field.__class__.__name__)()
            # Mark PK
            pk_suffix = " PK" if getattr(field, 'primary_key', False) else ""
            lines.append(f"        {field_type} {field_name}{pk_suffix}")
        lines.append("    }")

    # Collect relations from FK and O2O
    for model in all_models:
        for field in model._meta.get_fields():
            if getattr(field, 'many_to_many', False):
                # Many-to-many: A }o--o{ B
                try:
                    target = field.related_model
                    if target is None:
                        continue
                    a = model.__name__
                    b = target.__name__
                    lines.append(f"    {a} }}o--o{{ {b} : many_to_many")
                except Exception:
                    continue
            elif getattr(field, 'many_to_one', False) or getattr(field, 'one_to_one', False):
                # For FK (many-to-one): Parent ||--o{ Child
                try:
                    target = field.related_model
                    if target is None:
                        continue
                    parent = target.__name__
                    child = model.__name__
                    connector = "||--||" if getattr(field, 'one_to_one', False) else "||--o{"
                    rel_name = getattr(field, 'name', 'fk')
                    lines.append(f"    {parent} {connector} {child} : {rel_name}")
                except Exception:
                    continue

    return "\n".join(lines)


@login_required
@user_passes_test(is_admin)
def er_diagram_page(request):
    """Render a page that shows the ER diagram using Mermaid."""
    mermaid = _generate_mermaid_er_diagram()
    return render(request, 'dashboard/er.html', { 'mermaid': mermaid })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_er_diagram(request):
    """Return Mermaid ER diagram string."""
    mermaid = _generate_mermaid_er_diagram()
    return Response({ 'mermaid': mermaid })


@login_required
@user_passes_test(is_admin)
def test_openpyxl(request):
    """Test endpoint to check openpyxl import."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        return HttpResponse("openpyxl works fine!")
    except ImportError as e:
        return HttpResponse(f"ImportError: {e}", status=500)
    except Exception as e:
        return HttpResponse(f"Other error: {e}", status=500)


# ------------------------------
# Student Management Views
# ------------------------------
@login_required
@user_passes_test(is_admin)
def students_list(request):
    """Students management page"""
    from students.models import Student
    
    # Get filter parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    grade = request.GET.get('grade', '')
    section = request.GET.get('section', '')
    
    # Build queryset
    students = Student.objects.all()
    
    if search:
        students = students.filter(
            Q(roll_number__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(father_name__icontains=search) |
            Q(mother_name__icontains=search)
        )
    
    if status:
        students = students.filter(status=status)
    
    if grade:
        students = students.filter(grade_level=grade)
    
    if section:
        students = students.filter(section=section)
    
    students = students.order_by('-created_at')
    
    # Get grade choices for filter
    from students.models import Student
    grade_choices = Student.GRADE_CHOICES
    
    context = {
        'students': students,
        'grade_choices': grade_choices,
    }
    return render(request, 'dashboard/students.html', context)


@login_required
@user_passes_test(is_admin)
def student_detail(request, student_id):
    """Student detail page"""
    try:
        student = Student.objects.get(id=student_id)
        context = {
            'student': student,
            'enrollment_history': student.enrollment_history.all(),
            'documents': student.documents.all(),
            'custom_fields': student.custom_field_values.all(),
        }
        return render(request, 'dashboard/student_detail.html', context)
    except Student.DoesNotExist:
        return render(request, 'dashboard/404.html', status=404)


@login_required
@user_passes_test(is_admin)
def custom_fields_list(request):
    """Custom fields management page"""
    from students.models import CustomField
    from django.utils import timezone
    
    custom_fields = CustomField.objects.all().order_by('order', 'name')
    
    # Calculate statistics
    active_fields_count = custom_fields.filter(is_active=True).count()
    required_fields_count = custom_fields.filter(required=True).count()
    field_types_count = custom_fields.values('field_type').distinct().count()
    
    context = {
        'custom_fields': custom_fields,
        'active_fields_count': active_fields_count,
        'required_fields_count': required_fields_count,
        'field_types_count': field_types_count,
        'now': timezone.now(),
    }
    return render(request, 'dashboard/custom_fields.html', context)


@login_required
@user_passes_test(is_admin)
def student_login_page(request):
    """Student login page"""
    return render(request, 'dashboard/student_login.html')


@login_required
@user_passes_test(is_admin)
def student_sessions(request):
    """Student login sessions page"""
    from django.utils import timezone
    from datetime import timedelta
    
    # Get filter parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Build queryset
    sessions = UserSession.objects.filter(user__student_profile__isnull=False)
    
    if search:
        sessions = sessions.filter(
            Q(user__student_profile__roll_number__icontains=search) |
            Q(user__student_profile__first_name__icontains=search) |
            Q(user__student_profile__last_name__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    if status == 'active':
        sessions = sessions.filter(revoked=False, expires_at__gt=timezone.now())
    elif status == 'expired':
        sessions = sessions.filter(expires_at__lt=timezone.now())
    elif status == 'revoked':
        sessions = sessions.filter(revoked=True)
    
    if date_from:
        sessions = sessions.filter(created_at__date__gte=date_from)
    
    if date_to:
        sessions = sessions.filter(created_at__date__lte=date_to)
    
    sessions = sessions.order_by('-created_at')
    
    # Calculate statistics
    active_sessions_count = UserSession.objects.filter(
        user__student_profile__isnull=False,
        revoked=False,
        expires_at__gt=timezone.now()
    ).count()
    
    unique_students_count = UserSession.objects.filter(
        user__student_profile__isnull=False
    ).values('user').distinct().count()
    
    today = timezone.now().date()
    today_logins_count = UserSession.objects.filter(
        user__student_profile__isnull=False,
        created_at__date=today
    ).count()
    
    context = {
        'sessions': sessions,
        'active_sessions_count': active_sessions_count,
        'unique_students_count': unique_students_count,
        'today_logins_count': today_logins_count,
        'now': timezone.now(),
    }
    return render(request, 'dashboard/student_sessions.html', context)


# ------------------------------
# Student Import Views
# ------------------------------
@login_required
@user_passes_test(is_admin)
def student_import_page(request):
    """Student import page"""
    recent_imports = StudentImport.objects.all()[:10]
    context = {
        'recent_imports': recent_imports,
    }
    return render(request, 'dashboard/student_import.html', context)


@login_required
@user_passes_test(is_admin)
def student_import_process(request):
    """Process student import from Excel/CSV file"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'success': False, 'error': 'No file uploaded'})
        
        # Validate file type
        allowed_extensions = ['.xlsx', '.xls', '.csv']
        file_extension = os.path.splitext(file.name)[1].lower()
        if file_extension not in allowed_extensions:
            return JsonResponse({'success': False, 'error': 'Invalid file type. Please upload Excel or CSV file.'})
        
        # Get import options
        skip_errors = request.POST.get('skip_errors') == 'on'
        create_login = request.POST.get('create_login') == 'on'
        update_existing = request.POST.get('update_existing') == 'on'
        
        # Create import record
        import_record = StudentImport.objects.create(
            filename=file.name,
            file_size=file.size,
            created_by=request.user,
            skip_errors=skip_errors,
            create_login=create_login,
            update_existing=update_existing,
            status='PROCESSING'
        )
        
        # Process the file
        result = process_student_import(file, import_record, skip_errors, create_login, update_existing)
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@user_passes_test(is_admin)
def download_template(request):
    """Download Excel template for student import"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Create sample data
        sample_data = {
            'roll_number': ['STU001', 'STU002', 'STU003'],
            'first_name': ['John', 'Jane', 'Mike'],
            'last_name': ['Doe', 'Smith', 'Johnson'],
            'middle_name': ['A', 'B', 'C'],
            'date_of_birth': ['2010-01-15', '2010-03-20', '2010-06-10'],
            'gender': ['M', 'F', 'M'],
            'grade_level': ['10', '11', '12'],
            'section': ['A', 'B', 'A'],
            'academic_year': ['2023-2024', '2023-2024', '2023-2024'],
            'email': ['john.doe@example.com', 'jane.smith@example.com', 'mike.johnson@example.com'],
            'student_mobile': ['+1234567890', '+1234567891', '+1234567892'],
            'quota': ['GENERAL', 'SC', 'OBC'],
            'rank': ['1', '5', '10'],
            'status': ['ACTIVE', 'ACTIVE', 'ACTIVE'],
            'father_name': ['John Doe Sr', 'James Smith', 'Robert Johnson'],
            'mother_name': ['Mary Doe', 'Sarah Smith', 'Lisa Johnson'],
            'father_mobile': ['+1234567893', '+1234567894', '+1234567895'],
            'mother_mobile': ['+1234567896', '+1234567897', '+1234567898'],
            'address': ['123 Main St', '456 Oak Ave', '789 Pine Rd'],
            'city': ['New York', 'Los Angeles', 'Chicago'],
            'state': ['NY', 'CA', 'IL'],
            'country': ['USA', 'USA', 'USA'],
            'postal_code': ['10001', '90210', '60601'],
        }
        
        # Create DataFrame
        df = pd.DataFrame(sample_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Students']
            
            # Add instructions sheet
            instructions_data = {
                'Field': [
                    'roll_number', 'first_name', 'last_name', 'middle_name', 'date_of_birth', 'gender',
                    'grade_level', 'section', 'academic_year', 'email', 'student_mobile', 'quota', 'rank',
                    'status', 'father_name', 'mother_name', 'father_mobile', 'mother_mobile', 'address',
                    'city', 'state', 'country', 'postal_code'
                ],
                'Required': [
                    'Yes', 'Yes', 'Yes', 'No', 'Yes', 'Yes', 'No', 'No', 'No', 'No', 'No', 'No', 'No',
                    'No', 'No', 'No', 'No', 'No', 'No', 'No', 'No', 'No', 'No'
                ],
                'Description': [
                    'Unique student identifier', 'Student first name', 'Student last name', 'Student middle name',
                    'Date of birth (YYYY-MM-DD)', 'Gender: M (Male), F (Female), O (Other)',
                    'Grade level: 1-12', 'Section: A, B, C, D, E', 'Academic year (e.g., 2023-2024)',
                    'Student email address', 'Student phone number', 'Quota category', 'Academic rank',
                    'Status: ACTIVE, INACTIVE, GRADUATED', 'Father name', 'Mother name', 'Father phone',
                    'Mother phone', 'Full address', 'City', 'State', 'Country', 'Postal code'
                ],
                'Example': [
                    'STU001', 'John', 'Doe', 'A', '2010-01-15', 'M', '10', 'A', '2023-2024',
                    'john.doe@example.com', '+1234567890', 'GENERAL', '1', 'ACTIVE', 'John Doe Sr',
                    'Mary Doe', '+1234567893', '+1234567896', '123 Main St', 'New York', 'NY', 'USA', '10001'
                ]
            }
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
        
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        return response
        
    except ImportError:
        return JsonResponse({'error': 'openpyxl is not installed'}, status=500)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def process_student_import(file, import_record, skip_errors, create_login, update_existing):
    """Process student import from file"""
    try:
        import pandas as pd
        from datetime import datetime
        
        # Read file based on extension
        file_extension = os.path.splitext(file.name)[1].lower()
        
        if file_extension == '.csv':
            df = pd.read_csv(file)
        else:  # Excel files
            df = pd.read_excel(file)
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Initialize counters
        success_count = 0
        error_count = 0
        warning_count = 0
        errors = []
        warnings = []
        
        # Update import record
        import_record.total_rows = len(df)
        import_record.save()
        
        # Process each row
        for index, row in df.iterrows():
            row_number = index + 2  # +2 because Excel is 1-indexed and we have header
            
            try:
                # Validate required fields
                required_fields = ['roll_number', 'first_name', 'last_name', 'date_of_birth', 'gender']
                for field in required_fields:
                    if pd.isna(row.get(field, '')) or str(row[field]).strip() == '':
                        raise ValueError(f"Required field '{field}' is missing or empty")
                
                # Clean and validate data
                roll_number = str(row['roll_number']).strip()
                first_name = str(row['first_name']).strip()
                last_name = str(row['last_name']).strip()
                date_of_birth = str(row['date_of_birth']).strip()
                gender = str(row['gender']).strip().upper()
                
                # Handle optional grade_level
                grade_level = None
                if not pd.isna(row.get('grade_level', '')) and str(row['grade_level']).strip() != '':
                    grade_level = str(row['grade_level']).strip()
                
                # Validate gender
                if gender not in ['M', 'F', 'O']:
                    raise ValueError(f"Invalid gender '{gender}'. Must be M, F, or O")
                
                # Validate grade level (if provided)
                if grade_level is not None and grade_level not in [str(i) for i in range(1, 13)]:
                    raise ValueError(f"Invalid grade level '{grade_level}'. Must be 1-12")
                
                # Validate date format
                try:
                    datetime.strptime(date_of_birth, '%Y-%m-%d')
                except ValueError:
                    raise ValueError(f"Invalid date format '{date_of_birth}'. Use YYYY-MM-DD")
                
                # Check if student exists
                existing_student = None
                if update_existing:
                    existing_student = Student.objects.filter(roll_number=roll_number).first()
                
                if existing_student:
                    # Update existing student
                    student = existing_student
                    warnings.append({
                        'row': row_number,
                        'field': 'roll_number',
                        'message': f"Student with roll number '{roll_number}' already exists. Updating."
                    })
                    warning_count += 1
                else:
                    # Create new student
                    student = Student()
                
                # Set basic fields
                student.roll_number = roll_number
                student.first_name = first_name
                student.last_name = last_name
                student.date_of_birth = date_of_birth
                student.gender = gender
                if grade_level is not None:
                    student.grade_level = grade_level
                
                # Set optional fields
                if not pd.isna(row.get('middle_name', '')):
                    student.middle_name = str(row['middle_name']).strip()
                
                if not pd.isna(row.get('email', '')):
                    student.email = str(row['email']).strip()
                
                if not pd.isna(row.get('student_mobile', '')):
                    student.student_mobile = str(row['student_mobile']).strip()
                
                if not pd.isna(row.get('section', '')):
                    section = str(row['section']).strip().upper()
                    if section in ['A', 'B', 'C', 'D', 'E']:
                        student.section = section
                
                if not pd.isna(row.get('academic_year', '')):
                    student.academic_year = str(row['academic_year']).strip()
                
                if not pd.isna(row.get('quota', '')):
                    quota = str(row['quota']).strip().upper()
                    if quota in [choice[0] for choice in Student.QUOTA_CHOICES]:
                        student.quota = quota
                
                if not pd.isna(row.get('rank', '')):
                    try:
                        student.rank = int(row['rank'])
                    except (ValueError, TypeError):
                        pass
                
                if not pd.isna(row.get('status', '')):
                    status = str(row['status']).strip().upper()
                    if status in [choice[0] for choice in Student.STATUS_CHOICES]:
                        student.status = status
                
                # Set parent information
                if not pd.isna(row.get('father_name', '')):
                    student.father_name = str(row['father_name']).strip()
                
                if not pd.isna(row.get('mother_name', '')):
                    student.mother_name = str(row['mother_name']).strip()
                
                if not pd.isna(row.get('father_mobile', '')):
                    student.father_mobile = str(row['father_mobile']).strip()
                
                if not pd.isna(row.get('mother_mobile', '')):
                    student.mother_mobile = str(row['mother_mobile']).strip()
                
                # Set address information
                address_parts = []
                for field in ['address', 'city', 'state', 'country', 'postal_code']:
                    if not pd.isna(row.get(field, '')):
                        address_parts.append(str(row[field]).strip())
                
                if address_parts:
                    student.full_address = ', '.join(address_parts)
                
                # Save student
                student.save()
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append({
                    'row': row_number,
                    'field': 'general',
                    'message': str(e)
                })
                
                if not skip_errors:
                    break
        
        # Update import record
        import_record.success_count = success_count
        import_record.error_count = error_count
        import_record.warning_count = warning_count
        import_record.errors = errors
        import_record.warnings = warnings
        import_record.status = 'COMPLETED'
        import_record.save()
        
        return {
            'success': True,
            'success_count': success_count,
            'error_count': error_count,
            'warning_count': warning_count,
            'errors': errors,
            'warnings': warnings
        }
        
    except Exception as e:
        import_record.status = 'FAILED'
        import_record.errors = [{'row': 0, 'field': 'general', 'message': str(e)}]
        import_record.save()
        
        return {
            'success': False,
            'error': str(e)
        }

# API Testing Dashboard Views
@login_required
@user_passes_test(is_admin)
def api_testing_dashboard(request):
    """Main API testing dashboard"""
    context = {
        'total_collections': APICollection.objects.filter(created_by=request.user).count(),
        'total_requests': APIRequest.objects.filter(collection__created_by=request.user).count(),
        'total_tests': APITest.objects.filter(request__collection__created_by=request.user).count(),
        'total_environments': APIEnvironment.objects.filter(created_by=request.user).count(),
        'recent_results': APITestResult.objects.filter(
            test__request__collection__created_by=request.user
        ).order_by('-executed_at')[:10],
        'recent_suite_results': APITestSuiteResult.objects.filter(
            suite__collection__created_by=request.user
        ).order_by('-started_at')[:5],
        'active_automations': APIAutomation.objects.filter(
            test_suite__collection__created_by=request.user,
            is_active=True
        ).count(),
    }
    return render(request, 'dashboard/api_testing/dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def api_collections_list(request):
    """API Collections management page"""
    if request.method == 'POST':
        # Handle collection creation
        try:
            collection = APICollection.objects.create(
                name=request.POST.get('name'),
                description=request.POST.get('description', ''),
                base_url=request.POST.get('base_url', ''),
                is_public=request.POST.get('is_public') == 'on',
                created_by=request.user
            )
            return JsonResponse({
                'success': True,
                'message': 'Collection created successfully',
                'collection_id': str(collection.id)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    collections = APICollection.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, 'dashboard/api_testing/collections.html', {'collections': collections})

@login_required
@user_passes_test(is_admin)
def api_collection_detail_view(request, collection_id):
    """Get collection details for AJAX requests"""
    try:
        collection = APICollection.objects.get(id=collection_id, created_by=request.user)
        return JsonResponse({
            'id': str(collection.id),
            'name': collection.name,
            'description': collection.description,
            'base_url': collection.base_url,
            'is_public': collection.is_public,
            'created_at': collection.created_at.isoformat(),
            'updated_at': collection.updated_at.isoformat()
        })
    except APICollection.DoesNotExist:
        return JsonResponse({'error': 'Collection not found'}, status=404)

@login_required
@user_passes_test(is_admin)
def api_collection_update_view(request, collection_id):
    """Update collection for AJAX requests"""
    try:
        collection = APICollection.objects.get(id=collection_id, created_by=request.user)
        if request.method == 'PUT':
            # Parse PUT data
            import json
            data = json.loads(request.body.decode('utf-8'))
            
            collection.name = data.get('name', collection.name)
            collection.description = data.get('description', collection.description)
            collection.base_url = data.get('base_url', collection.base_url)
            collection.is_public = data.get('is_public', collection.is_public)
            collection.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Collection updated successfully'
            })
    except APICollection.DoesNotExist:
        return JsonResponse({'error': 'Collection not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_collection_duplicate_view(request, collection_id):
    """Duplicate collection for AJAX requests"""
    try:
        collection = APICollection.objects.get(id=collection_id, created_by=request.user)
        
        # Create new collection
        new_collection = APICollection.objects.create(
            name=f"{collection.name} (Copy)",
            description=collection.description,
            base_url=collection.base_url,
            is_public=collection.is_public,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Collection duplicated successfully',
            'new_collection_id': str(new_collection.id)
        })
    except APICollection.DoesNotExist:
        return JsonResponse({'error': 'Collection not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_collection_delete_view(request, collection_id):
    """Delete collection for AJAX requests"""
    try:
        collection = APICollection.objects.get(id=collection_id, created_by=request.user)
        collection.delete()
        return JsonResponse({
            'success': True,
            'message': 'Collection deleted successfully'
        })
    except APICollection.DoesNotExist:
        return JsonResponse({'error': 'Collection not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_collection_detail(request, collection_id):
    """API Collection detail page"""
    collection = get_object_or_404(APICollection, id=collection_id, created_by=request.user)
    return render(request, 'dashboard/api_testing/collection_detail.html', {'collection': collection})

@login_required
@user_passes_test(is_admin)
def api_environments_list(request):
    """API Environments management page"""
    if request.method == 'POST':
        # Handle environment creation
        try:
            import json
            variables = json.loads(request.POST.get('variables', '{}'))
            
            # If setting as default, unset other defaults
            if request.POST.get('is_default') == 'on':
                APIEnvironment.objects.filter(created_by=request.user, is_default=True).update(is_default=False)
            
            environment = APIEnvironment.objects.create(
                name=request.POST.get('name'),
                description=request.POST.get('description', ''),
                variables=variables,
                is_default=request.POST.get('is_default') == 'on',
                created_by=request.user
            )
            return JsonResponse({
                'success': True,
                'message': 'Environment created successfully',
                'environment_id': str(environment.id)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    environments = APIEnvironment.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, 'dashboard/api_testing/environments.html', {'environments': environments})

@login_required
@user_passes_test(is_admin)
def api_environment_detail_view(request, environment_id):
    """Get environment details for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        return JsonResponse({
            'id': str(environment.id),
            'name': environment.name,
            'description': environment.description,
            'variables': environment.variables,
            'is_default': environment.is_default,
            'created_at': environment.created_at.isoformat(),
            'updated_at': environment.updated_at.isoformat()
        })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)

@login_required
@user_passes_test(is_admin)
def api_environment_update_view(request, environment_id):
    """Update environment for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        if request.method == 'PUT':
            # Parse PUT data
            import json
            data = json.loads(request.body.decode('utf-8'))
            
            environment.name = data.get('name', environment.name)
            environment.description = data.get('description', environment.description)
            environment.variables = data.get('variables', environment.variables)
            environment.is_default = data.get('is_default', environment.is_default)
            
            # If setting as default, unset other defaults
            if environment.is_default:
                APIEnvironment.objects.filter(created_by=request.user, is_default=True).exclude(id=environment.id).update(is_default=False)
            
            environment.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Environment updated successfully'
            })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_environment_duplicate_view(request, environment_id):
    """Duplicate environment for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        
        # Create new environment
        new_environment = APIEnvironment.objects.create(
            name=f"{environment.name} (Copy)",
            description=environment.description,
            variables=environment.variables,
            is_default=False,  # Don't duplicate default status
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Environment duplicated successfully',
            'new_environment_id': str(new_environment.id)
        })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_environment_set_default_view(request, environment_id):
    """Set environment as default for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        
        # Unset other defaults
        APIEnvironment.objects.filter(created_by=request.user, is_default=True).update(is_default=False)
        
        # Set this environment as default
        environment.is_default = True
        environment.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Environment set as default successfully'
        })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_environment_delete_view(request, environment_id):
    """Delete environment for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        environment.delete()
        return JsonResponse({
            'success': True,
            'message': 'Environment deleted successfully'
        })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_requests_list(request):
    """API Requests management page"""
    requests = APIRequest.objects.filter(collection__created_by=request.user).order_by('collection__name', 'order')
    return render(request, 'dashboard/api_testing/requests.html', {'requests': requests})

@login_required
@user_passes_test(is_admin)
def api_request_detail(request, request_id):
    """API Request detail page"""
    api_request = get_object_or_404(APIRequest, id=request_id, collection__created_by=request.user)
    return render(request, 'dashboard/api_testing/request_detail.html', {'api_request': api_request})

@login_required
@user_passes_test(is_admin)
def api_tests_list(request):
    """API Tests management page"""
    tests = APITest.objects.filter(request__collection__created_by=request.user).order_by('request__collection__name', 'request__order', 'name')
    return render(request, 'dashboard/api_testing/tests.html', {'tests': tests})

@login_required
@user_passes_test(is_admin)
def api_test_detail(request, test_id):
    """API Test detail page"""
    test = get_object_or_404(APITest, id=test_id, request__collection__created_by=request.user)
    return render(request, 'dashboard/api_testing/test_detail.html', {'test': test})

@login_required
@user_passes_test(is_admin)
def api_test_results_list(request):
    """API Test Results page"""
    results = APITestResult.objects.filter(
        test__request__collection__created_by=request.user
    ).order_by('-executed_at')
    return render(request, 'dashboard/api_testing/test_results.html', {'results': results})

@login_required
@user_passes_test(is_admin)
def api_test_suites_list(request):
    """API Test Suites management page"""
    suites = APITestSuite.objects.filter(collection__created_by=request.user).order_by('-created_at')
    return render(request, 'dashboard/api_testing/test_suites.html', {'suites': suites})

@login_required
@user_passes_test(is_admin)
def api_test_suite_detail(request, suite_id):
    """API Test Suite detail page"""
    suite = get_object_or_404(APITestSuite, id=suite_id, collection__created_by=request.user)
    return render(request, 'dashboard/api_testing/test_suite_detail.html', {'suite': suite})

@login_required
@user_passes_test(is_admin)
def api_automations_list(request):
    """API Automations management page"""
    automations = APIAutomation.objects.filter(test_suite__collection__created_by=request.user).order_by('-created_at')
    return render(request, 'dashboard/api_testing/automations.html', {'automations': automations})

@login_required
@user_passes_test(is_admin)
def api_automation_detail(request, automation_id):
    """API Automation detail page"""
    automation = get_object_or_404(APIAutomation, id=automation_id, test_suite__collection__created_by=request.user)
    return render(request, 'dashboard/api_testing/automation_detail.html', {'automation': automation})

@login_required
@user_passes_test(is_admin)
def api_testing_workspace(request):
    """Interactive API testing workspace"""
    collections = APICollection.objects.filter(created_by=request.user).order_by('name')
    environments = APIEnvironment.objects.filter(created_by=request.user).order_by('name')
    return render(request, 'dashboard/api_testing/workspace.html', {
        'collections': collections,
        'environments': environments
    })

@user_passes_test(is_admin)
def simple_api_workspace(request):
    """Simple API testing workspace - Postman-like interface"""
    collections = APICollection.objects.filter(created_by=request.user).order_by('name')
    return render(request, 'dashboard/api_testing/simple_workspace.html', {
        'collections': collections
    })

# ------------------------------
# Faculty Management Views
# ------------------------------
@login_required
@user_passes_test(is_admin)
def faculty_list(request):
    """Faculty management page"""
    # Get filter parameters
    search = request.GET.get('search', '')
    department = request.GET.get('department', '')
    status = request.GET.get('status', '')
    
    # Build queryset
    faculties = Faculty.objects.all()
    
    if search:
        faculties = faculties.filter(
            Q(name__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(apaar_faculty_id__icontains=search) |
            Q(email__icontains=search) |
            Q(phone_number__icontains=search)
        )
    
    if department:
        faculties = faculties.filter(department=department)
    
    if status:
        faculties = faculties.filter(status=status)
    
    faculties = faculties.order_by('name')
    
    # Get statistics
    total_faculty = Faculty.objects.count()
    active_faculty = Faculty.objects.filter(status='ACTIVE', currently_associated=True).count()
    
    context = {
        'faculties': faculties,
        'total_faculty': total_faculty,
        'active_faculty': active_faculty,
        'department_choices': Faculty.DEPARTMENT_CHOICES,
        'status_choices': Faculty.STATUS_CHOICES,
    }
    return render(request, 'dashboard/faculty/list.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_detail(request, faculty_id):
    """Faculty detail page"""
    try:
        faculty = Faculty.objects.get(id=faculty_id)
        context = {
            'faculty': faculty,
            'subjects': faculty.subjects.all(),
            'schedules': faculty.schedules.all(),
            'leaves': faculty.leaves.all(),
            'performance_records': faculty.performance_records.all(),
            'documents': faculty.documents.all(),
            'custom_fields': faculty.custom_field_values.all(),
        }
        return render(request, 'dashboard/faculty/detail.html', context)
    except Faculty.DoesNotExist:
        return render(request, 'dashboard/404.html', status=404)

@login_required
@user_passes_test(is_admin)
def custom_fields_list(request):
    """Custom fields management page"""
    from students.models import CustomField
    from django.utils import timezone
    
    custom_fields = CustomField.objects.all().order_by('order', 'name')
    
    # Calculate statistics
    active_fields_count = custom_fields.filter(is_active=True).count()
    required_fields_count = custom_fields.filter(required=True).count()
    field_types_count = custom_fields.values('field_type').distinct().count()
    
    context = {
        'custom_fields': custom_fields,
        'active_fields_count': active_fields_count,
        'required_fields_count': required_fields_count,
        'field_types_count': field_types_count,
        'now': timezone.now(),
    }
    return render(request, 'dashboard/custom_fields.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_performance_stats(request):
    """Faculty performance statistics dashboard"""
    # Get total faculty count
    total_faculty = Faculty.objects.count()
    
    # Get performance statistics
    performance_stats = FacultyPerformance.objects.all().aggregate(
        total_performance_records=Count('id'),
        average_rating=Avg('overall_score')
    )
    
    # Get recent performance records
    recent_performance_records = FacultyPerformance.objects.order_by('-created_at')[:10]
    
    # Get department-wise statistics
    department_stats = Faculty.objects.values('department').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'total_faculty': total_faculty,
        'performance_stats': performance_stats,
        'recent_performance_records': recent_performance_records,
        'department_stats': department_stats,
    }
    return render(request, 'dashboard/faculty/performance_stats.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_leave_stats(request):
    """Faculty leave statistics dashboard"""
    # Get total faculty count
    total_faculty = Faculty.objects.count()
    
    # Get leave statistics
    leave_stats = FacultyLeave.objects.all().aggregate(
        total_leaves=Count('id'),
        approved_leaves=Count('id', filter=Q(status='APPROVED')),
        pending_leaves=Count('id', filter=Q(status='PENDING')),
        rejected_leaves=Count('id', filter=Q(status='REJECTED')),
    )
    
    # Get recent leaves
    recent_leaves = FacultyLeave.objects.order_by('-created_at')[:10]
    
    # Get leave type statistics
    leave_type_stats = FacultyLeave.objects.values('leave_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'total_faculty': total_faculty,
        'leave_stats': leave_stats,
        'recent_leaves': recent_leaves,
        'leave_type_stats': leave_type_stats,
    }
    return render(request, 'dashboard/faculty/leave_stats.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_document_list(request):
    """Faculty document management page"""
    documents = FacultyDocument.objects.all().order_by('-created_at')
    
    context = {
        'documents': documents,
    }
    return render(request, 'dashboard/faculty/documents.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_custom_field_create(request):
    """Create a new faculty custom field"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.POST.get('name')
            label = request.POST.get('label')
            field_type = request.POST.get('field_type')
            required = request.POST.get('required') == 'on'
            is_active = request.POST.get('is_active') == 'on'
            default_value = request.POST.get('default_value', '')
            choices = request.POST.get('choices', '')
            help_text = request.POST.get('help_text', '')
            order = request.POST.get('order', 0)
            
            # Validate required fields
            if not all([name, label, field_type]):
                return JsonResponse({'success': False, 'error': 'Name, label, and field type are required'})
            
            # Check if name already exists
            if FacultyCustomField.objects.filter(name=name).exists():
                return JsonResponse({'success': False, 'error': 'A field with this name already exists'})
            
            # Create the custom field
            custom_field = FacultyCustomField.objects.create(
                name=name,
                label=label,
                field_type=field_type,
                required=required,
                is_active=is_active,
                default_value=default_value,
                choices=choices,
                help_text=help_text,
                order=order
            )
            
            return JsonResponse({'success': True, 'message': 'Custom field created successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
@user_passes_test(is_admin)
def faculty_custom_field_update(request, field_id):
    """Update an existing faculty custom field"""
    if request.method == 'POST':
        try:
            # Get the custom field
            try:
                custom_field = FacultyCustomField.objects.get(id=field_id)
            except FacultyCustomField.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Custom field not found'})
            
            # Get form data
            name = request.POST.get('name')
            label = request.POST.get('label')
            field_type = request.POST.get('field_type')
            required = request.POST.get('required') == 'on'
            is_active = request.POST.get('is_active') == 'on'
            default_value = request.POST.get('default_value', '')
            choices = request.POST.get('choices', '')
            help_text = request.POST.get('help_text', '')
            order = request.POST.get('order', 0)
            
            # Validate required fields
            if not all([name, label, field_type]):
                return JsonResponse({'success': False, 'error': 'Name, label, and field type are required'})
            
            # Check if name already exists (excluding current field)
            if FacultyCustomField.objects.filter(name=name).exclude(id=field_id).exists():
                return JsonResponse({'success': False, 'error': 'A field with this name already exists'})
            
            # Update the custom field
            custom_field.name = name
            custom_field.label = label
            custom_field.field_type = field_type
            custom_field.required = required
            custom_field.is_active = is_active
            custom_field.default_value = default_value
            custom_field.choices = choices
            custom_field.help_text = help_text
            custom_field.order = order
            custom_field.save()
            
            return JsonResponse({'success': True, 'message': 'Custom field updated successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
@user_passes_test(is_admin)
def faculty_custom_field_delete(request, field_id):
    """Delete a faculty custom field"""
    if request.method == 'POST':
        try:
            # Get the custom field
            try:
                custom_field = FacultyCustomField.objects.get(id=field_id)
            except FacultyCustomField.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Custom field not found'})
            
            # Delete the custom field
            custom_field.delete()
            
            return JsonResponse({'success': True, 'message': 'Custom field deleted successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
@user_passes_test(is_admin)
def faculty_custom_fields_list(request):
    """Faculty custom fields management page"""
    custom_fields = FacultyCustomField.objects.all().order_by('order', 'name')
    
    # Calculate statistics
    active_fields_count = custom_fields.filter(is_active=True).count()
    required_fields_count = custom_fields.filter(required=True).count()
    field_types_count = custom_fields.values('field_type').distinct().count()
    
    context = {
        'custom_fields': custom_fields,
        'active_fields_count': active_fields_count,
        'required_fields_count': required_fields_count,
        'field_types_count': field_types_count,
        'now': timezone.now(),
    }
    return render(request, 'dashboard/faculty/custom_fields.html', context)